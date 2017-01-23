#!/usr/bin/env python
#-*- coding:utf-8 -*-
'''
目的: 把student.txt转换为xls格式的文档
'''

import re,string
from openpyxl import Workbook
def readfile(text):
        f=open(text,'r')
        lines=f.readlines()
        f.close()
        return lines

def txt_to_xlsx():
	line=readfile('student.txt')
	data=[]
	wb=Workbook()
	ws=wb.create_sheet(title='student')
	#用于写入到A1,A2,A3,A4,A5
	l=string.letters.upper()[:26]

	for i in line:
		if '{' not in i and '}' not in i:
			line_re=i.strip().strip('],')
#			print	line_re.replace(':[',',')
#			print	re.sub(':\[',',',line_re)
			data.append(re.sub(':\[',',',line_re).split(','))
	for a in range(len(data)):
		for b in range(len(data[a])):
			ws[l[b]+str(a+1)]=data[a][b].strip('"').strip()
	wb.save('student.xlsx')

if __name__ == '__main__':
	txt_to_xlsx()
